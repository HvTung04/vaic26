"""Generate GapLens contract Excel sheet — data models + API endpoints."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
section_font = Font(bold=True, size=12, color="1F4E79")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = thin_border

def style_rows(ws, start, end, cols):
    for r in range(start, end + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap

# ─── Sheet 1: Data Models ───
ws1 = wb.active
ws1.title = "Data Models"
cols1 = ["Table", "Column", "Type", "Nullable", "FK / Constraint", "Description"]
widths1 = [18, 22, 14, 10, 28, 45]
for i, (h, w) in enumerate(zip(cols1, widths1), 1):
    ws1.cell(row=1, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = w
style_header(ws1, 1, len(cols1))

models = [
    # User
    ("User", "id", "uuid PK", "No", "", "Primary key"),
    ("User", "role", "text", "No", "", '"student" | "teacher"'),
    ("User", "name", "text", "No", "", "Display name"),
    ("User", "email", "text", "Yes", "", "Optional email"),
    ("User", "class_id", "uuid FK", "Yes", "-> Class.id", "Student's class"),
    ("User", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # Class
    ("Class", "id", "uuid PK", "No", "", "Primary key"),
    ("Class", "name", "text", "No", "", 'e.g. "7A1"'),
    ("Class", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # KnowledgeNode
    ("KnowledgeNode", "id", "text PK", "No", "", '"math-g5-fraction-equivalent"'),
    ("KnowledgeNode", "label", "text", "No", "", "Vietnamese label"),
    ("KnowledgeNode", "subject", "text", "No", "", '"math"'),
    ("KnowledgeNode", "grade", "smallint", "No", "", "5, 6, 7..."),
    ("KnowledgeNode", "unit", "smallint", "No", "", "Current teaching unit (proximity)"),
    ("KnowledgeNode", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # Prerequisite
    ("Prerequisite", "node_id", "text FK", "No", "-> KnowledgeNode.id", "Dependent node"),
    ("Prerequisite", "prereq_id", "text FK", "No", "-> KnowledgeNode.id", "Prerequisite node"),
    # Question
    ("Question", "id", "uuid PK", "No", "", "Primary key"),
    ("Question", "text", "text", "No", "", "LaTeX preserved"),
    ("Question", "options", "jsonb", "No", "", '[{"key":"A","text":"..."}]'),
    ("Question", "correct_answer", "text", "No", "", '"B"'),
    ("Question", "knowledge_nodes", "jsonb", "No", "", '["node-a","node-b"] (multi)'),
    ("Question", "difficulty", "smallint", "No", "", "1=easy 2=medium 3=hard"),
    ("Question", "confidence", "real", "No", "", "0..1 LLM tag confidence"),
    ("Question", "source_type", "text", "No", "", '"pdf" | "photo"'),
    ("Question", "status", "text", "No", "", '"draft" | "approved"'),
    ("Question", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # Exam
    ("Exam", "id", "uuid PK", "No", "", "Primary key"),
    ("Exam", "teacher_id", "uuid FK", "No", "-> User.id", "Creator"),
    ("Exam", "title", "text", "No", "", "Exam title"),
    ("Exam", "type", "text", "No", "", '"weekly" | "revision"'),
    ("Exam", "question_ids", "jsonb", "No", "", "[uuid, uuid, ...] ordered"),
    ("Exam", "status", "text", "No", "", '"draft" | "active" | "completed"'),
    ("Exam", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # StudentAnswer
    ("StudentAnswer", "id", "uuid PK", "No", "", "Primary key"),
    ("StudentAnswer", "exam_id", "uuid FK", "No", "-> Exam.id", "Which exam"),
    ("StudentAnswer", "student_id", "uuid FK", "No", "-> User.id", "Which student"),
    ("StudentAnswer", "question_id", "uuid FK", "No", "-> Question.id", "Which question"),
    ("StudentAnswer", "selected_answer", "text", "No", "", '"A"|"B"|"C"|"D"'),
    ("StudentAnswer", "is_correct", "boolean", "No", "", "Derived from correct_answer"),
    ("StudentAnswer", "time_spent_sec", "real", "No", "", "Seconds to answer"),
    ("StudentAnswer", "confidence", "real", "Yes", "", "0..1 student self-report (app)"),
    ("StudentAnswer", "hesitation", "real", "Yes", "", "Time before first tap (app)"),
    ("StudentAnswer", "source", "text", "No", "", '"app" | "bubble_sheet"'),
    ("StudentAnswer", "answered_at", "timestamptz", "No", "", "Answer timestamp"),
    # MasteryRecord
    ("MasteryRecord", "id", "uuid PK", "No", "", "Primary key"),
    ("MasteryRecord", "student_id", "uuid FK", "No", "-> User.id", "Which student"),
    ("MasteryRecord", "node_id", "text FK", "No", "-> KnowledgeNode.id", "Which knowledge node"),
    ("MasteryRecord", "mastery_level", "real", "No", "", "0..1"),
    ("MasteryRecord", "weight", "real", "No", "", "difficulty_weight x unit_proximity"),
    ("MasteryRecord", "confidence", "real", "No", "", "Model certainty"),
    ("MasteryRecord", "updated_at", "timestamptz", "No", "", "Last update timestamp"),
    # BubbleSheet
    ("BubbleSheet", "id", "uuid PK", "No", "", "Primary key"),
    ("BubbleSheet", "exam_id", "uuid FK", "No", "-> Exam.id", "Linked exam"),
    ("BubbleSheet", "generated_pdf", "bytea", "Yes", "", "PDF bytes or file path"),
    ("BubbleSheet", "created_at", "timestamptz", "No", "", "Creation timestamp"),
    # BubbleAnswer
    ("BubbleAnswer", "id", "uuid PK", "No", "", "Primary key"),
    ("BubbleAnswer", "bubble_sheet_id", "uuid FK", "No", "-> BubbleSheet.id", "Which sheet"),
    ("BubbleAnswer", "question_index", "smallint", "No", "", "Row number on sheet"),
    ("BubbleAnswer", "question_id", "uuid FK", "Yes", "-> Question.id", "Resolved question"),
    ("BubbleAnswer", "detected_option", "text", "No", "", '"A"|"B"|"C"|"D"'),
    ("BubbleAnswer", "confidence", "real", "No", "", "Detection confidence"),
]

row = 2
current_table = None
start_row = 2
for table, col, typ, nullable, fk, desc in models:
    if table != current_table:
        if current_table is not None:
            style_rows(ws1, start_row, row - 1, len(cols1))
            row += 1  # blank row between tables
        style_section(ws1, row, len(cols1), table)
        row += 1
        start_row = row
        current_table = table
    ws1.cell(row=row, column=1, value=table)
    ws1.cell(row=row, column=2, value=col)
    ws1.cell(row=row, column=3, value=typ)
    ws1.cell(row=row, column=4, value=nullable)
    ws1.cell(row=row, column=5, value=fk)
    ws1.cell(row=row, column=6, value=desc)
    row += 1
style_rows(ws1, start_row, row - 1, len(cols1))

# ─── Sheet 2: API Endpoints ───
ws2 = wb.create_sheet("API Endpoints")
cols2 = ["Group", "Method", "Path", "Body / Params", "Response", "Owned By", "Description"]
widths2 = [14, 10, 30, 40, 30, 12, 40]
for i, (h, w) in enumerate(zip(cols2, widths2), 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header(ws2, 1, len(cols2))

endpoints = [
    # Auth
    ("Auth", "POST", "/auth/login", "{ role, name }", "{ user }", "Role 3", "Fake login, seeded accounts"),
    ("Auth", "GET", "/auth/me", "", "{ user }", "Role 3", "Current user from session/token"),
    # Questions
    ("Questions", "POST", "/questions", "[{ text, options, correct_answer, knowledge_nodes, difficulty, ... }]", "[question] (draft)", "Role 1", "Bulk insert drafts"),
    ("Questions", "GET", "/questions", "?node=&difficulty=&status=", "[question]", "Role 1", "List/filter questions"),
    ("Questions", "GET", "/questions/{id}", "", "question", "Role 1", "Single question detail"),
    ("Questions", "PATCH", "/questions/{id}", "{ ...fields, status='approved' }", "question", "Role 1", "Teacher review: edit + approve"),
    ("Questions", "DELETE", "/questions/{id}", "", "204", "Role 1", "Remove draft"),
    # Exams
    ("Exams", "POST", "/exams", "{ title, type, question_ids }", "exam", "Role 3", "Create exam from approved questions"),
    ("Exams", "GET", "/exams", "", "[exam]", "Role 3", "List exams (teacher: all, student: assigned)"),
    ("Exams", "GET", "/exams/{id}", "", "exam + questions", "Role 3", "Exam detail with questions"),
    ("Exams", "PATCH", "/exams/{id}", "{ status }", "exam", "Role 3", "Status update (draft->active->completed)"),
    # Student Answers
    ("Answers", "POST", "/answers", "{ exam_id, answers: [{ question_id, selected_answer, time_spent_sec, confidence?, hesitation? }] }", "[answer]", "Role 3", "Submit answers (app or bubble sheet)"),
    ("Answers", "GET", "/answers", "?exam_id=&student_id=", "[answer]", "Role 3", "Query answers"),
    # Mastery
    ("Mastery", "GET", "/mastery/{student_id}", "", "[{ node_id, mastery_level, weight, confidence }]", "Role 2", "Student mastery per node"),
    ("Mastery", "GET", "/mastery/{student_id}/graph", "", "graph state", "Role 2", "Full graph for learning path"),
    ("Mastery", "POST", "/mastery/update", "{ exam_id }", "200", "Role 2", "Trigger recalc after answers (internal)"),
    # Dashboard
    ("Dashboard", "GET", "/dashboard/priority-queue", "", "[{ student, urgency, reasons }]", "Role 2", "Students needing help sorted by urgency"),
    ("Dashboard", "GET", "/dashboard/gap-radar", "", "[{ node_id, weak_count, total }]", "Role 2", "Class-wide weak nodes (radar chart)"),
    ("Dashboard", "GET", "/dashboard/groups", "", "[{ nodes, student_ids }]", "Role 2", "Need-based student clusters"),
    ("Dashboard", "GET", "/dashboard/intervention", "", "[{ type, target, reason }]", "Role 2", "Suggested actions"),
    # Bubble Sheet
    ("Bubble Sheet", "POST", "/bubble-sheets/generate", "{ exam_id }", "PDF bytes", "Role 1", "Generate printable bubble sheet"),
    ("Bubble Sheet", "POST", "/bubble-sheets/scan", "image file", "[{ question_index, detected_option, confidence }]", "Role 1", "Upload scan -> detect answers (preview)"),
    ("Bubble Sheet", "POST", "/bubble-sheets/confirm", "{ bubble_sheet_id, mappings: [{ question_id, detected_option }] }", "200", "Role 1", "Confirm scan -> create StudentAnswers"),
]

row = 2
current_group = None
start_row = 2
for group, method, path, body, resp, owner, desc in endpoints:
    if group != current_group:
        if current_group is not None:
            style_rows(ws2, start_row, row - 1, len(cols2))
            row += 1
        style_section(ws2, row, len(cols2), group)
        row += 1
        start_row = row
        current_group = group
    ws2.cell(row=row, column=1, value=group)
    ws2.cell(row=row, column=2, value=method)
    ws2.cell(row=row, column=3, value=path)
    ws2.cell(row=row, column=4, value=body)
    ws2.cell(row=row, column=5, value=resp)
    ws2.cell(row=row, column=6, value=owner)
    ws2.cell(row=row, column=7, value=desc)
    row += 1
style_rows(ws2, start_row, row - 1, len(cols2))

# ─── Sheet 3: Answer Metrics ───
ws3 = wb.create_sheet("Answer Metrics")
cols3 = ["Metric", "Column", "Source", "Type", "When Null", "Why It Matters"]
widths3 = [18, 22, 16, 12, 18, 45]
for i, (h, w) in enumerate(zip(cols3, widths3), 1):
    ws3.cell(row=1, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header(ws3, 1, len(cols3))

metrics = [
    ("Correctness", "is_correct", "derived", "boolean", "never", "Basic scoring: correct_answer == selected_answer"),
    ("Time Spent", "time_spent_sec", "app timer", "real", "never", "Fast-right != slow-right (guess vs mastery)"),
    ("Confidence", "confidence", "student self-report", "real 0..1", "bubble sheet", "Student-assessed certainty of answer"),
    ("Hesitation", "hesitation", "app (first tap delay)", "real sec", "bubble sheet", "Uncertainty signal before interaction"),
    ("Source", "source", "system", "text", "never", '"app" vs "bubble_sheet" — context for metric reliability'),
    ("Selected Answer", "selected_answer", "app / scan", "text", "never", "Raw response before grading"),
    ("Difficulty Weight", "weight", "Role 2 formula", "real", "pre-answer", "difficulty_weight[1|2|3] × unit_proximity"),
    ("Unit Proximity", "(computed)", "graph distance", "real 0..1", "pre-answer", "1.0 = current unit, decays by prerequisite distance"),
]

for i, (metric, col, source, typ, null, why) in enumerate(metrics, 2):
    ws3.cell(row=i, column=1, value=metric)
    ws3.cell(row=i, column=2, value=col)
    ws3.cell(row=i, column=3, value=source)
    ws3.cell(row=i, column=4, value=typ)
    ws3.cell(row=i, column=5, value=null)
    ws3.cell(row=i, column=6, value=why)
style_rows(ws3, 2, len(metrics) + 1, len(cols3))

# ─── Sheet 4: Mastery Formula ───
ws4 = wb.create_sheet("Mastery Formula")
ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 60

formula_data = [
    ("Mastery Update Formula", ""),
    ("", ""),
    ("weight(node)", "= difficulty_weight × unit_proximity"),
    ("", ""),
    ("difficulty_weight[1] (easy)", "1.0"),
    ("difficulty_weight[2] (medium)", "1.5"),
    ("difficulty_weight[3] (hard)", "2.0"),
    ("", ""),
    ("unit_proximity", "1.0 if node == current teaching unit"),
    ("unit_proximity", "decays by graph distance (prerequisite hops)"),
    ("unit_proximity", "far-away prerequisite => low proximity => weak signal"),
    ("", ""),
    ("Wrong answer on hard, nearby node", "=> strong mastery drop"),
    ("Right answer on easy, far-away node", "=> weak mastery gain"),
    ("", ""),
    ("Root-cause diagnosis", "trace prerequisites: wrong node X + parent mastery < threshold => suspect parent"),
    ("Confidence threshold", "if root-cause confidence < threshold => state uncertainty, safer suggestion"),
    ("Revision test selector", "rule-based: 2-3 weakest nodes, easy -> hard"),
    ("Learning path", "graph state -> LLM generates explanation + review order"),
]

style_section(ws4, 1, 2, "Mastery Formula & Root-Cause Logic")
for i, (k, v) in enumerate(formula_data, 3):
    ws4.cell(row=i, column=1, value=k).font = Font(bold=True) if k else Font()
    ws4.cell(row=i, column=2, value=v)

# Save
out = Path(__file__).resolve().parents[1] / "contract.xlsx"
wb.save(str(out))
print(f"saved: {out}")
